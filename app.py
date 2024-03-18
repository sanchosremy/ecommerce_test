import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    feuil = wb['Feuil1']
    cell= feuil['a1']
    cell = feuil.cell(1, 1)

    for row in range(2, feuil.max_row + 1):
        cell = feuil.cell(row, 3)
        corrected_price = cell.value * 0.9
        print(corrected_price)
        corrected_price_sell = feuil.cell(row, 4)
        corrected_price_sell.value = corrected_price

    values = Reference(feuil, 
            min_row=2, 
            max_row=feuil.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    feuil.add_chart(chart, 'e2')

    wb.save(filename)